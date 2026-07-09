from __future__ import annotations

import re
from pathlib import Path
from typing import Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NUMERIC_HEADER_TERMS = {
    "count",
    "rows",
    "row",
    "gsm",
    "gse",
    "unique",
    "duplicate",
    "duplicated",
    "seconds",
    "minutes",
    "runtime",
    "percent",
    "pct",
    "value",
    "n_",
    "size",
}

TEXT_WRAP_HEADER_TERMS = {
    "path",
    "reason",
    "explanation",
    "json",
    "info",
    "summary",
    "message",
    "rationale",
    "evidence",
    "description",
    "action",
    "notes",
    "url",
}


def safe_excel_sheet_name(name: str, used: set[str] | None = None) -> str:
    used = used if used is not None else set()
    base = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(name).strip())[:31] or "Sheet"
    out = base
    i = 1

    while out in used:
        suffix = f"_{i}"
        out = base[: 31 - len(suffix)] + suffix
        i += 1

    used.add(out)
    return out


def _looks_numeric_header(header: str) -> bool:
    h = str(header).strip().lower()
    return any(term in h for term in NUMERIC_HEADER_TERMS)


def _looks_wrappable_header(header: str) -> bool:
    h = str(header).strip().lower()
    return any(term in h for term in TEXT_WRAP_HEADER_TERMS)


def _status_fill(value: str) -> PatternFill | None:
    v = str(value).strip().upper()

    if v in {"PASS", "INFO", "INCLUDE", "TRUE"}:
        return PatternFill("solid", fgColor="E2F0D9")

    if v in {"REVIEW", "WARNING", "RETRY", "PENDING"} or "REVIEW" in v:
        return PatternFill("solid", fgColor="FFF2CC")

    if v in {"FAIL", "ERROR", "FALSE", "EXCLUDE"} or "FAIL" in v:
        return PatternFill("solid", fgColor="FCE4D6")

    return None


def format_excel_workbook(path: str | Path) -> Path:
    """
    Apply professional report formatting to an existing .xlsx workbook.

    Formatting policy:
    - bold centered headers
    - freeze first row
    - enable autofilter
    - center numeric/status columns
    - left-align and wrap long text columns
    - set sensible column widths
    """
    path = Path(path)

    if not path.exists() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return path

    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        ws.freeze_panes = "A2"

        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

        headers = []
        for cell in ws[1]:
            headers.append("" if cell.value is None else str(cell.value))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for col_idx in range(1, ws.max_column + 1):
            header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ""
            letter = get_column_letter(col_idx)

            max_len = len(str(header))
            numeric_like = _looks_numeric_header(header)
            wrap_like = _looks_wrappable_header(header)
            status_like = str(header).strip().lower() in {
                "status",
                "global_qc_status",
                "review_required",
                "release_action",
                "mapping_qa_status",
                "release_ready",
            }

            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value

                if value is not None:
                    max_len = max(max_len, min(len(str(value)), 80))

                cell.border = border

                if status_like:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    fill = _status_fill(value)
                    if fill is not None:
                        cell.fill = fill

                elif numeric_like or isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

                elif wrap_like:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

            if wrap_like:
                ws.column_dimensions[letter].width = min(max(max_len + 2, 18), 60)
            elif numeric_like or status_like:
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 22)
            else:
                ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 36)

        ws.row_dimensions[1].height = 24

    wb.save(path)
    return path


def write_dataframes_to_excel(
    path: str | Path,
    sheets: Mapping[str, pd.DataFrame],
    *,
    format_workbook: bool = True,
) -> Path:
    """
    Write a dict of DataFrames to Excel and apply standard formatting.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    used: set[str] = set()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = safe_excel_sheet_name(sheet_name, used)
            if df is None:
                df = pd.DataFrame()
            df.to_excel(writer, sheet_name=safe_name, index=False)

    if format_workbook:
        format_excel_workbook(path)

    return path